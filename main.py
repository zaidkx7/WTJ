import os
import json
import openpyxl
import requests

from bs4 import BeautifulSoup
from logger import get_logger
from json.decoder import JSONDecodeError
from openpyxl.styles import PatternFill, Alignment


__all__ = ['WTJScrapper']


class WTJScrapper:
    BASE_URL = "https://www.welcometothejungle.com/en/companies"
    API_URL = "https://api.welcometothejungle.com/api/v1/pages?path=/en/companies/"
    COMPANY_URL_TEMPLATE = "https://www.welcometothejungle.com/en/companies/{slug}"
    PROJECT_DIR = os.path.dirname(os.path.abspath(__file__))
    RESPONSE_DIR = os.path.join(PROJECT_DIR, 'response')
    if not os.path.exists(RESPONSE_DIR):
        os.makedirs(RESPONSE_DIR)
    SESSION = requests.Session()
    COMPANIES_INFO = []
    SECTOR_MAP = {}

    def __init__(self):
        self.logger = get_logger()
        self.logger.info("Initializing WTJScrapper...")

    def get_page(self, url):
        response = self.SESSION.get(url)
        response.raise_for_status()
        return response.text

    def extract_json_data(self, soup):
        script_tag = soup.find(
            'script', string=lambda s: s and 'window.__INITIAL_DATA__' in s)
        if script_tag:
            script_content = script_tag.text
            try:
                json_data = script_content.split('window.__INITIAL_DATA__ = ')[1]
                json_data = json_data.split('window.__GROWTHBOOK_PAYLOAD__ = ')[0]
                return json.loads(json_data)
            except (IndexError, json.JSONDecodeError) as e:
                self.logger.error(f"Error parsing JSON data: {e}")
                return None
        return None

    def get_necessary_json_data(self, json_data):
        try:
            parsed_data = json.loads(json_data)
            return parsed_data['queries'][2]['state']['data']['results']
        except (KeyError, IndexError, json.JSONDecodeError) as e:
            self.logger.error(f"Error extracting necessary data: {e}")
            return None

    def get_company_slugs(self, json_data):
        slugs = []

        for data in json_data:
            if 'hits' in data:
                for company in data['hits']:
                    slug = company.get('slug', None)
                    if slug:
                        slugs.append(slug)
                        self.SECTOR_MAP[slug] = company.get('sectors', [])
        return slugs

    def gather_api_calls(self, company_slugs):
        return [self.API_URL + slug for slug in company_slugs]

    def get_company_website(self, slug):
        url = self.COMPANY_URL_TEMPLATE.format(slug=slug)
        page = self.get_page(url)
        soup = BeautifulSoup(page, 'html.parser')
        website_tag = soup.find(
            'a', class_='sc-fyVfxW hXemWC sc-eHsDsR hnvrnA')
        if website_tag and website_tag['href']:
            return website_tag['href']

    def extract_company_data(self, api_calls):
        for call in api_calls:
            response = self.SESSION.get(call)
            response.raise_for_status()
            response_json = response.json()

            sections = response_json['page']['sections']
            company_info = {
                'name': None,
                'location': None,
                'url': None,
                'website': None,
                'sectors': None,
                'social-networks': None,
                'description': None,
                'presentation': None,
                'what_they_are_looking_for': None,
                'good_to_know': None,
            }

            for section in sections:
                for container in section.get('containers', []):
                    for block in container.get('blocks', []):
                        for content in block.get('contents', []):
                            properties = content.get('properties', {})

                            if properties.get('organization'):
                                company_info['name'] = properties.get(
                                    'organization', {}).get('name', None)

                            if content.get('kind') == 'map':
                                company_info['location'] = properties.get(
                                    'headquarter', {}).get('city', None)

                            if content.get('kind') == 'company-stats':
                                company_info[content.get('kind')] = properties

                            if content.get('kind') == 'social-networks':
                                company_info['social-networks'] = properties.get(
                                    'networks', {})

                            if content.get('kind') == 'text':
                                if properties.get('title') == 'Good to know' or properties.get('title') == 'Bon à savoir':
                                    company_info['good_to_know'] = properties.get(
                                        'body', None)
                                elif properties.get('title') == 'What they are looking for' or properties.get('title') == "Ce qu'ils recherchent":
                                    company_info['what_they_are_looking_for'] = properties.get(
                                        'body', None)
                                elif properties.get('title') == 'Presentation' or properties.get('title') == 'Présentation':
                                    company_info['presentation'] = properties.get(
                                        'body', None)

            company_slug = call.split('/')[-1]

            company_info['sectors'] = self.SECTOR_MAP.get(company_slug, [])
            company_info['website'] = self.get_company_website(company_slug)
            company_info['url'] = self.COMPANY_URL_TEMPLATE.format(
                slug=company_slug)
            company_info['description'] = response_json['page']['metas']['description']

            self.COMPANIES_INFO.append(company_info)

        return self.COMPANIES_INFO

    def save_to_json(self, json_data):
        with open(f'{self.RESPONSE_DIR}/data.json', 'w') as file:
            json.dump(json_data, file, indent=4)

    def save_to_excel(self, data):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Companies Info"

        headers = [
            "id", "Name", "Location", "Website", "URL", "Sectors", "Facebook", "Linkedin", "Twitter", "Youtube", "Description",
            "Presentation", "What They Are Looking For", "Good To Know"
        ]

        header_fill = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True)

        id = 0
        for row_num, company in enumerate(data, start=2):
            row_data = [
                id,
                company.get('name'),
                company.get('location'),
                company.get('website'),
                company.get('url'),
                ", ".join(company.get('sectors', [])),
                company.get('social-networks', {}).get('facebook', None),
                company.get('social-networks', {}).get('linkedin', None),
                company.get('social-networks', {}).get('twitter', None),
                company.get('social-networks', {}).get('youtube', None),
                company.get('description'),
                company.get('presentation'),
                company.get('what_they_are_looking_for'),
                company.get('good_to_know')
            ]
            id += 1
            for col_num, value in enumerate(row_data, start=1):
                cell = sheet.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(
                    horizontal="left", vertical="top", wrap_text=True)

        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            if max_length > 25:
                max_length = 25
            adjusted_width = max_length + 1.5
            sheet.column_dimensions[column_letter].width = adjusted_width

        for row in sheet.iter_rows():
            row_height = 15
            sheet.row_dimensions[row[0].row].height = row_height

        output_file = f"{self.RESPONSE_DIR}/companies_info.xlsx"
        workbook.save(output_file)
        self.logger.info(f"Companies info saved to Excel: {output_file}")

    def run(self):
        self.logger.info("Starting to scrape WTJ")
        page = self.get_page(self.BASE_URL)
        soup = BeautifulSoup(page, 'html.parser')
        self.logger.info('WTJ page scraped')
        try:
            self.logger.info('Trying to extract raw JSON from page')
            json_data = self.extract_json_data(soup)
            self.logger.info('Got raw JSON')
            self.logger.info('Trying to extract necessary data from JSON')
            json_data = self.get_necessary_json_data(json_data)
            self.logger.info('Got necessary data')
            self.save_to_json(json_data)
            if json_data:
                self.logger.info('Getting company slugs')
                slugs = self.get_company_slugs(json_data)
                self.logger.info('Got company slugs')
                api_calls = self.gather_api_calls(slugs)
                self.logger.info('Gathered API calls')
                companies = self.extract_company_data(api_calls)
                self.logger.info('Extracted company data')
                self.save_to_json(companies)
                self.logger.info('Companies info saved to JSON')
        except JSONDecodeError as e:
            self.logger.error(f"Error extracting JSON data: {e}")
            return None

def run():
    return WTJScrapper()


if __name__ == "__main__":
    run().run()
